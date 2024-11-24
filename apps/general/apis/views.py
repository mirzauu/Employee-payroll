import openpyxl
from openpyxl.reader.excel import load_workbook
from rest_framework.views import APIView
from rest_framework import generics
from .serializers import ServiceSerializer, DesignationSerializer, ServiceOptionSerializer, BankSerializer, \
    BankOptionSerializer, DesignationOptionSerializer, UploadFileSerializer
from apps.general.models import Services, Designation, Banks
from apps.clients.apis.views import CustomPagination
from apps.clients.mixins import OptionMixin
from rest_framework.permissions import IsAdminUser
from rest_framework.parsers import FileUploadParser
from django.http import HttpResponse
from openpyxl import Workbook
from django.template.loader import get_template

class ServiceListView(generics.ListCreateAPIView):
    permission_classes = [IsAdminUser]
    queryset = Services.objects.all()
    serializer_class = ServiceSerializer
    pagination_class = CustomPagination


class ServiceUpdateView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAdminUser]
    queryset = Services.objects.all()
    serializer_class = ServiceSerializer
    lookup_field = 'id'


class ServiceOptionView(generics.ListAPIView):
    permission_classes = [IsAdminUser]
    queryset = Services.objects.all()
    serializer_class = ServiceOptionSerializer


class DesignationView(generics.ListCreateAPIView):
    permission_classes = [IsAdminUser]
    queryset = Designation.objects.all()
    serializer_class = DesignationSerializer
    pagination_class = CustomPagination


class DesignationDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAdminUser]
    queryset = Designation.objects.all()
    serializer_class = DesignationSerializer
    lookup_field = 'id'


class DesignationOptionView(OptionMixin, generics.GenericAPIView):
    permission_classes = [IsAdminUser]
    queryset = Designation.objects.all()
    serializer_class = DesignationOptionSerializer


class ClientDesignation(OptionMixin, generics.GenericAPIView):

    queryset = Designation.objects.all()
    serializer_class = DesignationOptionSerializer
    lookup_field = 'id'

    def get_queryset(self):
        queryset = Designation.objects.filter(client_designations=self.kwargs['id'])
        return queryset


class BankOptionView(OptionMixin, generics.GenericAPIView):
    permission_classes = [IsAdminUser]
    queryset = Banks.objects.all()
    serializer_class = BankOptionSerializer


class BankView(generics.CreateAPIView):
    permission_classes = [IsAdminUser]
    queryset = Banks.objects.all()
    serializer_class = BankSerializer


class UploadFileView(generics.GenericAPIView):

    def match_template(self, template_path, template_name):
        workbook = load_workbook(template_path)
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f"attachment; filename={template_name}"
        workbook.save(response)
        return response

    def get(self, request, *args, **kwargs):
        option = request.GET.get('file_type', None)
        workbook = Workbook()
        # match option:
        #     case 'employee':
        #         template_path = 'templates/excel_template/employee.xlsx'
        #         employee_xl = self.match_template(template_path, 'employee')
        #         return employee_xl
        #     case 'bank':
        #         template_path = 'templates/excel_template/emp_bank.xlsx'
        #         employee_bank = self.match_template(template_path, 'employee_bank')
        #         return employee_bank
        #     case _:
        #         raise TypeError("No Type")
        pass


class DemoView(generics.GenericAPIView):

    serializer_class = UploadFileSerializer
    parser_classes = [FileUploadParser]

    # def post(self, request):
    #     file = request.data['file']
    #     if file.name.split('.')[-1] == 'xlsx':
    #         wb = openpyxl.load_workbook(file)
    #         sheet = wb.active
    #         for row in sheet.iter_rows(min_row=3, values_only=True):
    #             print(row)
    #             _data = {
    #                 "name": row[0]
    #             }
    #             obj = Test(**_data)
    #             obj.save()
    #         return Response({"Data Imported Succefully"})
    #     else:
    #         return Response({"message": "Invalid File Format"})


class DownloadExcel(APIView):

    def get(self, request, *args, **kwargs):
        pass
